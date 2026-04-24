import csv
import io
import json
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError as PydValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.persona import Persona
from app.schemas.import_result import ImportResult, ImportRowError
from app.schemas.persona import PersonaCreate

TEMPLATE_COLUMNS = [
    "name",
    "tone",
    "personality",
    "goal",
    "prompt_instructions",
    "constraints_json",
]


def _coerce_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Normalize CSV/XLSX row into PersonaCreate-compatible dict."""
    # strip all string values; treat empty strings as None
    def norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip()
            return v or None
        return v

    row = {k: norm(v) for k, v in raw.items() if k in TEMPLATE_COLUMNS}

    # Parse constraints_json
    constraints_raw = row.pop("constraints_json", None)
    if constraints_raw:
        if not isinstance(constraints_raw, str):
            constraints_raw = str(constraints_raw)
        try:
            constraints = json.loads(constraints_raw)
            if not isinstance(constraints, dict):
                raise ValueError("constraints_json must decode to a JSON object")
        except Exception as e:
            raise ValueError(f"invalid constraints_json: {e}")
        row["constraints"] = constraints
    else:
        row["constraints"] = {}

    if not row.get("name"):
        raise ValueError("name is required")

    return row


def _iter_rows_from_csv(data: bytes) -> Iterable[dict[str, Any]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for r in reader:
        yield r


def _iter_rows_from_xlsx(data: bytes) -> Iterable[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return
    columns = [str(h).strip() if h else "" for h in header]
    for values in rows_iter:
        if values is None:
            continue
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        yield {columns[i]: values[i] for i in range(min(len(columns), len(values)))}


class PersonaImportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_file(
        self, *, filename: str, content: bytes
    ) -> ImportResult:
        lower = filename.lower()
        if lower.endswith(".csv"):
            rows = list(_iter_rows_from_csv(content))
        elif lower.endswith(".xlsx"):
            rows = list(_iter_rows_from_xlsx(content))
        else:
            raise ValueError(
                "Unsupported file type. Use .csv or .xlsx."
            )

        errors: list[ImportRowError] = []
        to_create: list[Persona] = []
        for idx, raw in enumerate(rows, start=2):  # start=2 → first data row after header
            try:
                coerced = _coerce_row(raw)
                payload = PersonaCreate.model_validate(coerced)
                to_create.append(Persona(**payload.model_dump()))
            except PydValidationError as ve:
                errors.append(
                    ImportRowError(row=idx, message=str(ve.errors()[0]["msg"]))
                )
            except Exception as e:
                errors.append(ImportRowError(row=idx, message=str(e)))

        if to_create:
            self.db.add_all(to_create)
            await self.db.commit()

        return ImportResult(
            created=len(to_create),
            skipped=len(errors),
            errors=errors,
        )

    @staticmethod
    def build_template_xlsx() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "personas"
        ws.append(TEMPLATE_COLUMNS)
        ws.append(
            [
                "Maria — MVA victim",
                "distressed but cooperative",
                "Rear-ended on I-95; worried about bills.",
                "Find out if the firm can take her case.",
                "Volunteer PII after 45s if not asked.",
                '{"patience_sec": 60, "insurance": "GEICO"}',
            ]
        )
        ws.append(
            [
                "Robert — Slip & Fall",
                "frustrated",
                "Fell at Publix 3 weeks ago; impatient.",
                "Get confirmation the firm will take his case.",
                "Interrupt if agent monologues >10s.",
                '{"location": "Publix on Main St"}',
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
